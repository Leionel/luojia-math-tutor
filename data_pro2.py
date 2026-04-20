import json
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# --- 配置参数 ---
INPUT_JSON_FILE = "luojia_openmath_local_corpus_final.json"
OUTPUT_DIR = "xlsx_outputs"
RECORDS_PER_FILE = 10000

def clean_text_for_excel(text):
    """
    清理文本中的非法字符，使其能被 openpyxl 写入 Excel。
    移除控制字符和某些特殊的 LaTeX 命令残留。
    """
    if not isinstance(text, str):
        return str(text)
    
    # 移除 ASCII 控制字符（除了制表符、换行符）
    text = ''.join(char for char in text if ord(char) >= 32 or char in '\t\n\r')
    
    # 移除或替换某些 LaTeX 残留字符
    text = text.replace('\x00', '')  # 空字符
    text = text.replace('\x01', '')  # 其他控制字符
    text = re.sub(r'\\begin\{[^}]*\}', '[LATEX_BEGIN]', text)
    text = re.sub(r'\\end\{[^}]*\}', '[LATEX_END]', text)
    
    return text.strip()

def split_json_to_xlsx(json_file, output_dir, records_per_file):
    """
    按照每 records_per_file 条记录切割 JSON 文件，并保存到 XLSX。
    每个 XLSX 文件包含两列：问题（Question）和答案（Solution）
    
    参数:
        json_file: 输入的 JSON 文件路径
        output_dir: 输出目录
        records_per_file: 每个文件的记录数
    """
    # 创建输出目录
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
    
    # 读取 JSON 文件
    print(f"读取 JSON 文件：{json_file}")
    with open(json_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    total_records = len(data)
    print(f"总记录数：{total_records}")
    
    # 计算需要创建多少个文件
    num_files = (total_records + records_per_file - 1) // records_per_file
    
    # 按照 records_per_file 进行切割
    for file_index in range(num_files):
        start_idx = file_index * records_per_file
        end_idx = min(start_idx + records_per_file, total_records)
        
        # 提取当前批次的数据
        batch_data = data[start_idx:end_idx]
        
        # 创建 XLSX 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # 添加表头
        ws['A1'] = "Question"
        ws['B1'] = "Solution"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 80
        ws.column_dimensions['B'].width = 80
        
        # 添加数据行
        for row_idx, record in enumerate(batch_data, start=2):
            question = clean_text_for_excel(record.get('question', ''))
            solution = clean_text_for_excel(record.get('solution_with_code', ''))
            
            ws[f'A{row_idx}'] = question
            ws[f'B{row_idx}'] = solution
            
            # 设置行高自动换行
            ws[f'A{row_idx}'].alignment = Alignment(
                wrap_text=True, 
                vertical='top'
            )
            ws[f'B{row_idx}'].alignment = Alignment(
                wrap_text=True, 
                vertical='top'
            )
        
        # 生成输出文件名
        output_file = os.path.join(
            output_dir, 
            f"luojia_openmath_part_{file_index + 1:03d}_{start_idx}-{end_idx - 1}.xlsx"
        )
        
        # 保存文件
        wb.save(output_file)
        print(f"✓ 已保存：{output_file} (记录 {start_idx} - {end_idx - 1})")
    
    print(f"\n✓ 所有文件已生成！共 {num_files} 个文件")


if __name__ == "__main__":
    split_json_to_xlsx(
        json_file=INPUT_JSON_FILE,
        output_dir=OUTPUT_DIR,
        records_per_file=RECORDS_PER_FILE
    )